"""Exporta una DeclaracionMetodo a un .xlsx con el mismo formato real que usa
el cliente para sus Declaraciones de Método (5 hojas: Declaración de Método,
Firmas/Permisos/EPP, Catálogo de Peligros, Evaluación según Kinney y Control
del Documento) — mismos campos y mismas hojas de referencia que sus propios
ejemplos en Excel, para que el archivo se pueda usar tal cual en su proceso
sin tener que reformatearlo a mano."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .catalogo_peligros import CATALOGO_PELIGROS
from .models import nivel_riesgo

AZUL_OSCURO = "1F2A44"
GRIS_CLARO = "F3F4F6"
BORDE_FINO = Side(style="thin", color="D1D5DB")
BORDE_CELDA = Border(left=BORDE_FINO, right=BORDE_FINO, top=BORDE_FINO, bottom=BORDE_FINO)

TITULO_FONT = Font(bold=True, size=14, color="FFFFFF")
SUBTITULO_FONT = Font(bold=True, size=11)
ENCABEZADO_FONT = Font(bold=True, color="FFFFFF")
ENCABEZADO_FILL = PatternFill("solid", fgColor=AZUL_OSCURO)
ENVOLVER = Alignment(wrap_text=True, vertical="top")


def _titulo(hoja, texto, fila, col_ini, col_fin):
    hoja.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    celda = hoja.cell(row=fila, column=col_ini, value=texto)
    celda.font = TITULO_FONT
    celda.fill = ENCABEZADO_FILL
    celda.alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[fila].height = 22


def _etiqueta_valor(hoja, fila, col, etiqueta, valor):
    celda_etiqueta = hoja.cell(row=fila, column=col, value=etiqueta)
    celda_etiqueta.font = Font(bold=True)
    hoja.cell(row=fila, column=col + 1, value=valor or "")


def _encabezado_tabla(hoja, fila, encabezados, col_ini=1):
    for offset, texto in enumerate(encabezados):
        celda = hoja.cell(row=fila, column=col_ini + offset, value=texto)
        celda.font = ENCABEZADO_FONT
        celda.fill = ENCABEZADO_FILL
        celda.alignment = ENVOLVER
        celda.border = BORDE_CELDA


def _hoja_declaracion(hoja, declaracion, actividades):
    hoja.title = "Declaración de Método"
    fila = 1
    _titulo(hoja, "DECLARACIÓN DE MÉTODO Y EVALUACIÓN DE RIESGO", fila, 1, 15)
    fila += 2

    _etiqueta_valor(hoja, fila, 1, "Empresa contratista:", declaracion.contratista.nombre)
    _etiqueta_valor(hoja, fila, 5, "Planta / área:", declaracion.planta_area)
    fila += 1
    _etiqueta_valor(hoja, fila, 1, "Gerente de proyecto:", declaracion.gerente_proyecto)
    _etiqueta_valor(hoja, fila, 5, "Número de pedido:", declaracion.numero_pedido)
    fila += 1
    _etiqueta_valor(hoja, fila, 1, "Contacto:", declaracion.contacto_nombre)
    _etiqueta_valor(hoja, fila, 5, "Teléfono:", declaracion.contacto_telefono)
    fila += 1
    _etiqueta_valor(hoja, fila, 1, "Fecha de elaboración:", declaracion.fecha_elaboracion.isoformat())
    _etiqueta_valor(hoja, fila, 5, "Duración (días):", declaracion.duracion_dias)
    fila += 2

    celda = hoja.cell(row=fila, column=1, value="Describa el trabajo a realizar:")
    celda.font = Font(bold=True)
    fila += 1
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=15)
    hoja.cell(row=fila, column=1, value=declaracion.descripcion_trabajo).alignment = ENVOLVER
    fila += 2

    encabezados = [
        "Secuencia de actividades",
        "Técnicas / herramientas / equipos",
        "Descripción del riesgo",
        "P (sin)",
        "F (sin)",
        "I (sin)",
        "R (sin)",
        "Nivel (sin)",
        "Medidas de mitigación",
        "P (con)",
        "F (con)",
        "I (con)",
        "R (con)",
        "Nivel (con)",
        "Permisos de trabajo requeridos",
        "EPP requerido",
        "Tarea SIF",
    ]
    _encabezado_tabla(hoja, fila, encabezados)
    fila_tabla_inicio = fila + 1
    fila += 1

    for actividad in actividades:
        _, nivel_sin = nivel_riesgo(actividad["riesgo_sin"])
        _, nivel_con = nivel_riesgo(actividad["riesgo_con"])
        valores = [
            actividad["secuencia"],
            actividad["tecnicas_herramientas"],
            actividad["descripcion_riesgo"],
            actividad["probabilidad_sin"],
            actividad["frecuencia_sin"],
            actividad["impacto_sin"],
            actividad["riesgo_sin"],
            nivel_sin,
            actividad["medidas_mitigacion"],
            actividad["probabilidad_con"],
            actividad["frecuencia_con"],
            actividad["impacto_con"],
            actividad["riesgo_con"],
            nivel_con,
            ", ".join(actividad["permisos_requeridos"]),
            ", ".join(actividad["epp_requerido"]),
            "Sí" if actividad["tarea_sif"] else "No",
        ]
        for offset, valor in enumerate(valores):
            celda = hoja.cell(row=fila, column=1 + offset, value=valor)
            celda.alignment = ENVOLVER
            celda.border = BORDE_CELDA
        fila += 1

    anchos = [28, 26, 30, 7, 7, 7, 7, 12, 30, 7, 7, 7, 7, 12, 26, 22, 9]
    for indice, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho
    hoja.freeze_panes = hoja.cell(row=fila_tabla_inicio, column=1)


def _hoja_firmas_permisos_epp(hoja, declaracion, permisos_marcados, epp_marcado):
    hoja.title = "Firmas,Permisos, EPP"
    fila = 1
    _titulo(hoja, "FIRMAS, PERMISOS DE TRABAJO Y EPP", fila, 1, 6)
    fila += 2

    _encabezado_tabla(hoja, fila, ["Rol", "Nombre de quien firma", "Cuenta", "Firmado en"])
    fila += 1
    for firma in declaracion.firmas.all():
        valores = [
            firma.get_rol_display(),
            firma.nombre_firmante,
            firma.firmante_usuario.username if firma.firmante_usuario else "",
            firma.firmado_en.strftime("%Y-%m-%d %H:%M") if firma.firmado_en else "",
        ]
        for offset, valor in enumerate(valores):
            celda = hoja.cell(row=fila, column=1 + offset, value=valor)
            celda.border = BORDE_CELDA
        fila += 1
    if not declaracion.firmas.exists():
        hoja.cell(row=fila, column=1, value="Sin firmas registradas todavía.")
        fila += 1
    fila += 2

    celda = hoja.cell(row=fila, column=1, value="Permisos de trabajo / certificado de apoyo requerido")
    celda.font = SUBTITULO_FONT
    fila += 1
    _encabezado_tabla(hoja, fila, ["Permiso", "Requerido"], col_ini=1)
    fila += 1
    for nombre in permisos_marcados:
        hoja.cell(row=fila, column=1, value=nombre).border = BORDE_CELDA
        hoja.cell(row=fila, column=2, value="X").border = BORDE_CELDA
        fila += 1
    fila += 2

    celda = hoja.cell(row=fila, column=1, value="EPP a utilizar")
    celda.font = SUBTITULO_FONT
    fila += 1
    _encabezado_tabla(hoja, fila, ["EPP", "Requerido"], col_ini=1)
    fila += 1
    for nombre in epp_marcado:
        hoja.cell(row=fila, column=1, value=nombre).border = BORDE_CELDA
        hoja.cell(row=fila, column=2, value="X").border = BORDE_CELDA
        fila += 1

    hoja.column_dimensions["A"].width = 45
    hoja.column_dimensions["B"].width = 22
    hoja.column_dimensions["C"].width = 18
    hoja.column_dimensions["D"].width = 18


def _hoja_catalogo_peligros(hoja):
    hoja.title = "Catálogo de Peligros"
    encabezados = ["Tipo de peligro", "Materia / Energía", "Tipo de materia o energía", "Peligro", "Tipo de peligro", "Riesgo"]
    _encabezado_tabla(hoja, 1, encabezados)
    for fila, registro in enumerate(CATALOGO_PELIGROS, start=2):
        for offset, valor in enumerate(registro):
            celda = hoja.cell(row=fila, column=1 + offset, value=valor)
            celda.alignment = ENVOLVER
            celda.border = BORDE_CELDA
    anchos = [16, 16, 22, 32, 45, 40]
    for indice, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho


PROBABILIDADES = [
    (10, "Esperado"),
    (6, "Muy posible"),
    (3, "Raro"),
    (1, "Improbable pero posible"),
    (0.5, "Concebible pero improbable"),
    (0.1, "Casi improbable"),
]
FRECUENCIAS = [
    (10, "Continuamente", ""),
    (6, "Regularmente", "Diario"),
    (3, "De vez en cuando", "Semanalmente"),
    (2, "Algunas veces", "Mensualmente"),
    (1, "Rara vez", "Anual"),
    (0.5, "Muy rara vez", "Menos que una vez al año"),
]
IMPACTOS = [
    (40, "Catástrofe", "Varias fatalidades"),
    (15, "Muy serio", "Una fatalidad"),
    (7, "Serio", "Discapacidad"),
    (3, "Importante", "Lesión con baja"),
    (1, "Menor", "Lesión sin baja"),
]
RANGOS_RIESGO = [
    ("Más de 400", "Riesgo muy alto", "Detener esta actividad específica"),
    ("200 - 400", "Riesgo alto", "Requiere acción inmediata"),
    ("70 - 200", "Riesgo considerable", "Requiere corrección"),
    ("20 - 70", "Riesgo posible", "Requiere supervisión / atención"),
    ("Menos de 20", "Riesgo bajo", "Riesgo aceptable"),
]


def _hoja_kinney(hoja):
    hoja.title = "Evaluación según Kinney"
    celda = hoja.cell(row=1, column=1, value="Método de Kinney — R = Probabilidad (P) × Frecuencia (F) × Impacto (I)")
    celda.font = SUBTITULO_FONT
    fila = 3

    celda = hoja.cell(row=fila, column=1, value="Probabilidad (P)")
    celda.font = SUBTITULO_FONT
    fila += 1
    _encabezado_tabla(hoja, fila, ["Valor", "Si el evento es"])
    fila += 1
    for valor, descripcion in PROBABILIDADES:
        hoja.cell(row=fila, column=1, value=valor).border = BORDE_CELDA
        hoja.cell(row=fila, column=2, value=descripcion).border = BORDE_CELDA
        fila += 1
    fila += 1

    celda = hoja.cell(row=fila, column=1, value="Frecuencia de exposición (F)")
    celda.font = SUBTITULO_FONT
    fila += 1
    _encabezado_tabla(hoja, fila, ["Valor", "Si la acción/evento ocurre", "Lo cual significa"])
    fila += 1
    for valor, descripcion, significado in FRECUENCIAS:
        hoja.cell(row=fila, column=1, value=valor).border = BORDE_CELDA
        hoja.cell(row=fila, column=2, value=descripcion).border = BORDE_CELDA
        hoja.cell(row=fila, column=3, value=significado).border = BORDE_CELDA
        fila += 1
    fila += 1

    celda = hoja.cell(row=fila, column=1, value="Impacto (I)")
    celda.font = SUBTITULO_FONT
    fila += 1
    _encabezado_tabla(hoja, fila, ["Valor", "Impacto", "Lo cual significa"])
    fila += 1
    for valor, descripcion, significado in IMPACTOS:
        hoja.cell(row=fila, column=1, value=valor).border = BORDE_CELDA
        hoja.cell(row=fila, column=2, value=descripcion).border = BORDE_CELDA
        hoja.cell(row=fila, column=3, value=significado).border = BORDE_CELDA
        fila += 1
    fila += 1

    celda = hoja.cell(row=fila, column=1, value="Riesgo (R) — según el puntaje, tome las siguientes acciones")
    celda.font = SUBTITULO_FONT
    fila += 1
    _encabezado_tabla(hoja, fila, ["Puntaje", "Resultado de la evaluación", "Acción"])
    fila += 1
    for puntaje, resultado, accion in RANGOS_RIESGO:
        hoja.cell(row=fila, column=1, value=puntaje).border = BORDE_CELDA
        hoja.cell(row=fila, column=2, value=resultado).border = BORDE_CELDA
        hoja.cell(row=fila, column=3, value=accion).border = BORDE_CELDA
        fila += 1

    hoja.column_dimensions["A"].width = 16
    hoja.column_dimensions["B"].width = 26
    hoja.column_dimensions["C"].width = 30


def _hoja_control_documento(hoja, declaracion):
    hoja.title = "Control del Documento"
    from django.utils import timezone

    _etiqueta_valor(hoja, 1, 1, "Documento:", "Declaración de Método y Evaluación de Riesgo")
    _etiqueta_valor(hoja, 2, 1, "Declaración N°:", declaracion.pk)
    _etiqueta_valor(hoja, 3, 1, "Empresa contratista:", declaracion.contratista.nombre)
    _etiqueta_valor(hoja, 4, 1, "Estado:", declaracion.get_estado_display())
    _etiqueta_valor(hoja, 5, 1, "Fecha de elaboración:", declaracion.fecha_elaboracion.isoformat())
    _etiqueta_valor(hoja, 6, 1, "Generado el:", timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M"))
    _etiqueta_valor(hoja, 7, 1, "Generado por:", "SST Bavaria — Cámaras IA")
    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 40


def generar_excel_declaracion(declaracion):
    """Arma el libro completo — recibe la declaración con sus relaciones ya
    precargadas (contratista, actividades, firmas)."""
    actividades = []
    for actividad in declaracion.actividades.all():
        actividades.append(
            {
                "secuencia": actividad.secuencia,
                "tecnicas_herramientas": actividad.tecnicas_herramientas,
                "descripcion_riesgo": actividad.descripcion_riesgo,
                "probabilidad_sin": actividad.probabilidad_sin,
                "frecuencia_sin": actividad.frecuencia_sin,
                "impacto_sin": actividad.impacto_sin,
                "riesgo_sin": actividad.riesgo_sin,
                "medidas_mitigacion": actividad.medidas_mitigacion,
                "probabilidad_con": actividad.probabilidad_con,
                "frecuencia_con": actividad.frecuencia_con,
                "impacto_con": actividad.impacto_con,
                "riesgo_con": actividad.riesgo_con,
                "permisos_requeridos": actividad.permisos_requeridos,
                "epp_requerido": actividad.epp_requerido,
                "tarea_sif": actividad.tarea_sif,
            }
        )

    permisos_marcados = sorted({p for a in actividades for p in a["permisos_requeridos"]})
    epp_marcado = sorted({e for a in actividades for e in a["epp_requerido"]})

    libro = Workbook()
    _hoja_declaracion(libro.active, declaracion, actividades)
    _hoja_firmas_permisos_epp(libro.create_sheet(), declaracion, permisos_marcados, epp_marcado)
    _hoja_catalogo_peligros(libro.create_sheet())
    _hoja_kinney(libro.create_sheet())
    _hoja_control_documento(libro.create_sheet(), declaracion)
    return libro
