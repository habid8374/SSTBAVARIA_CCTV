"""Importa una Declaración de Método desde el Excel real que ya usa el
cliente (mismas 5 hojas que genera contratistas/exportar_declaracion_excel.py,
porque ese formato se construyó copiando este) — para que un contratista o
el personal de SST no tenga que retipear a mano un documento que ya existe.

Solo lee datos — nunca crea ni modifica nada en la base. Devuelve un dict
listo para precargar el formulario de Declaración de Método en el frontend;
la persona sigue revisando y guardando desde ahí, igual que si lo hubiera
escrito a mano. No importa contratista (lo decide quien sube el archivo, en
el propio formulario) ni firmas (una firma solo se puede crear firmando de
verdad, ligada a la cuenta autenticada — ver FirmaMetodo)."""

import re
import unicodedata

import openpyxl

from .models import EquipoProteccionPersonal, PermisoTrabajo


class ErrorImportacionExcel(Exception):
    """Mensaje pensado para mostrarse tal cual al usuario."""


HOJA_DECLARACION = "Declaración de Método"
HOJA_FIRMAS_PERMISOS_EPP = "Firmas,Permisos, EPP"

FILA_ENCABEZADO_TABLA = 12
FILA_PRIMERA_ACTIVIDAD = 14
COL_SECUENCIA = 1
COL_TECNICAS = 2
COL_DESCRIPCION_RIESGO = 3
COL_P_SIN, COL_F_SIN, COL_I_SIN = 4, 5, 6
COL_MEDIDAS = 8
COL_P_CON, COL_F_CON, COL_I_CON = 9, 10, 11
COL_PERMISOS = 13
COL_SIF = 14


def _normalizar(texto):
    """minúsculas, sin acentos, sin lo que venga entre paréntesis, sin
    puntuación colgante — para poder comparar el texto libre del Excel
    contra los nombres de los catálogos configurables."""
    texto = texto.split("(")[0]
    texto = texto.strip().rstrip(".,").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c))


def _emparejar_catalogo(etiqueta_cruda, nombres_catalogo):
    """Busca en el catálogo (ya normalizado) el nombre que mejor corresponda
    al texto libre del Excel. None si no encuentra nada razonable."""
    if not etiqueta_cruda or not etiqueta_cruda.strip():
        return None
    normalizada = _normalizar(etiqueta_cruda)
    if not normalizada:
        return None
    for nombre_original, nombre_normalizado in nombres_catalogo:
        if normalizada == nombre_normalizado:
            return nombre_original
    for nombre_original, nombre_normalizado in nombres_catalogo:
        if nombre_normalizado and (nombre_normalizado in normalizada or normalizada in nombre_normalizado):
            return nombre_original
    return None


def _celda_marcada(valor):
    """True si la celda de al lado de un ítem de catálogo trae 'X' (marcado)
    — cualquier otra cosa (vacío, 'NA', etc.) cuenta como no marcado."""
    return isinstance(valor, str) and valor.strip().upper() == "X"


def _extraer_regex(texto, patron, grupo=1):
    if not texto:
        return None
    coincidencia = re.search(patron, texto, re.IGNORECASE)
    return coincidencia.group(grupo).strip() if coincidencia else None


def _resolver_merge(hoja, fila, columna, mapa_merges):
    """Valor efectivo de una celda que puede estar dentro de una celda
    combinada — openpyxl solo guarda el valor en la esquina superior
    izquierda del rango combinado."""
    rango = mapa_merges.get((fila, columna))
    if rango is None:
        return hoja.cell(row=fila, column=columna).value
    return hoja.cell(row=rango.min_row, column=rango.min_col).value


def _es_inicio_de_merge(fila, columna, mapa_merges):
    rango = mapa_merges.get((fila, columna))
    return rango is None or (rango.min_row == fila and rango.min_col == columna)


def parsear_excel_declaracion(archivo):
    """`archivo` es un objeto tipo archivo (UploadedFile de Django). Devuelve
    un dict con los mismos campos que NuevaDeclaracion (sin `contratista` ni
    `estado`, que los define quien sube el archivo en el formulario) más
    `avisos`: una lista de textos sobre datos que no se pudieron reconocer
    automáticamente, para que la persona los revise antes de guardar."""
    try:
        libro = openpyxl.load_workbook(archivo, data_only=True)
    except Exception as exc:
        raise ErrorImportacionExcel("No se pudo abrir el archivo — asegúrate de que sea un .xlsx válido.") from exc

    if HOJA_DECLARACION not in libro.sheetnames:
        raise ErrorImportacionExcel(
            f'No se encontró la hoja "{HOJA_DECLARACION}" — este archivo no tiene el formato de '
            "Declaración de Método esperado."
        )
    hoja = libro[HOJA_DECLARACION]

    avisos = []

    planta_area = (hoja["B3"].value or "").strip()

    bloque_contacto = hoja["A6"].value or ""
    gerente_proyecto = _extraer_regex(bloque_contacto, r"GERENTE DE PROYECTO:\s*(.+)") or ""
    contacto_telefono = _extraer_regex(bloque_contacto, r"TEL[ÉE]FONO:\s*(.+)") or ""
    numero_pedido = _extraer_regex(bloque_contacto, r"N[ÚU]MERO DE PEDIDO:\s*(.+)") or ""

    bloque_fecha = hoja["C7"].value or ""
    fecha_texto = _extraer_regex(bloque_fecha, r"FECHA DE ELABORACI[ÓO]N:\s*(\d{1,2}/\d{1,2}/\d{2,4})")
    fecha_elaboracion = None
    if fecha_texto:
        dia, mes, anio = fecha_texto.split("/")
        if len(anio) == 2:
            anio = f"20{anio}"
        fecha_elaboracion = f"{anio}-{int(mes):02d}-{int(dia):02d}"
    else:
        avisos.append("No se pudo leer la fecha de elaboración — revísala antes de guardar.")

    duracion_texto = _extraer_regex(bloque_fecha, r"DURACI[ÓO]N[^:]*:\s*(\d+)")
    duracion_dias = int(duracion_texto) if duracion_texto else 1

    descripcion_trabajo = _extraer_regex(hoja["H7"].value or "", r"DESCRIBA EL TRABAJO A REALIZAR:\s*(.+)") or ""

    # --- Catálogo agregado de permisos/EPP marcados, de la hoja Firmas,Permisos,EPP ---
    permisos_marcados, epp_marcados = [], []
    if HOJA_FIRMAS_PERMISOS_EPP in libro.sheetnames:
        hoja_fpe = libro[HOJA_FIRMAS_PERMISOS_EPP]
        permisos_catalogo = [(p.nombre, _normalizar(p.nombre)) for p in PermisoTrabajo.objects.filter(activo=True)]
        epp_catalogo = [(e.nombre, _normalizar(e.nombre)) for e in EquipoProteccionPersonal.objects.filter(activo=True)]

        sin_mapear_permisos, sin_mapear_epp = set(), set()
        for fila in range(1, hoja_fpe.max_row + 1):
            etiqueta = hoja_fpe.cell(row=fila, column=9).value  # columna I
            if not etiqueta:
                continue
            marcada = hoja_fpe.cell(row=fila, column=11).value  # columna K
            if not _celda_marcada(marcada):
                continue
            emparejado = _emparejar_catalogo(str(etiqueta), permisos_catalogo)
            if emparejado:
                permisos_marcados.append(emparejado)
            else:
                sin_mapear_permisos.add(str(etiqueta).strip())

        for columna_etiqueta, columna_marca in (("L", "N"), ("O", "Q")):
            for fila in range(1, hoja_fpe.max_row + 1):
                etiqueta = hoja_fpe[f"{columna_etiqueta}{fila}"].value
                if not etiqueta:
                    continue
                marcada = hoja_fpe[f"{columna_marca}{fila}"].value
                if not _celda_marcada(marcada):
                    continue
                emparejado = _emparejar_catalogo(str(etiqueta), epp_catalogo)
                if emparejado:
                    epp_marcados.append(emparejado)
                else:
                    sin_mapear_epp.add(str(etiqueta).strip())

        permisos_marcados = sorted(set(permisos_marcados))
        epp_marcados = sorted(set(epp_marcados))
        if sin_mapear_permisos:
            avisos.append(
                "No reconocimos estos permisos marcados en el Excel — agrégalos a mano si aplican: "
                + ", ".join(sorted(sin_mapear_permisos))
            )
        if sin_mapear_epp:
            avisos.append(
                "No reconocimos este EPP marcado en el Excel — agrégalo a mano si aplica: "
                + ", ".join(sorted(sin_mapear_epp))
            )
    else:
        avisos.append(
            f'No se encontró la hoja "{HOJA_FIRMAS_PERMISOS_EPP}" — no se pudieron traer los '
            "permisos ni el EPP marcados; agrégalos a mano."
        )

    # --- Tabla de actividades, con celdas combinadas de "Secuencia" cubriendo varias filas de riesgo ---
    mapa_merges = {}
    for rango in hoja.merged_cells.ranges:
        for fila in range(rango.min_row, rango.max_row + 1):
            for columna in range(rango.min_col, rango.max_col + 1):
                mapa_merges[(fila, columna)] = rango

    actividades = []
    secuencia_actual = tecnicas_actual = ""
    orden = 0
    fila = FILA_PRIMERA_ACTIVIDAD
    filas_vacias_seguidas = 0
    while fila <= hoja.max_row and filas_vacias_seguidas < 3:
        if not _es_inicio_de_merge(fila, COL_DESCRIPCION_RIESGO, mapa_merges):
            fila += 1
            continue

        if _es_inicio_de_merge(fila, COL_SECUENCIA, mapa_merges):
            valor_secuencia = hoja.cell(row=fila, column=COL_SECUENCIA).value
            if valor_secuencia:
                secuencia_actual = str(valor_secuencia).strip()
                tecnicas_actual = str(_resolver_merge(hoja, fila, COL_TECNICAS, mapa_merges) or "").strip()

        descripcion_riesgo = hoja.cell(row=fila, column=COL_DESCRIPCION_RIESGO).value
        valores_numericos = [
            hoja.cell(row=fila, column=c).value
            for c in (COL_P_SIN, COL_F_SIN, COL_I_SIN, COL_P_CON, COL_F_CON, COL_I_CON)
        ]
        if not descripcion_riesgo and not any(v not in (None, "") for v in valores_numericos):
            filas_vacias_seguidas += 1
            fila += 1
            continue
        filas_vacias_seguidas = 0

        def _numero(columna):
            valor = hoja.cell(row=fila, column=columna).value
            try:
                return float(valor)
            except (TypeError, ValueError):
                return 0

        actividades.append(
            {
                "orden": orden,
                "secuencia": secuencia_actual,
                "tecnicas_herramientas": tecnicas_actual,
                "descripcion_riesgo": str(descripcion_riesgo or "").strip(),
                "probabilidad_sin": _numero(COL_P_SIN),
                "frecuencia_sin": _numero(COL_F_SIN),
                "impacto_sin": _numero(COL_I_SIN),
                "medidas_mitigacion": str(hoja.cell(row=fila, column=COL_MEDIDAS).value or "").strip(),
                "probabilidad_con": _numero(COL_P_CON),
                "frecuencia_con": _numero(COL_F_CON),
                "impacto_con": _numero(COL_I_CON),
                "permisos_requeridos": list(permisos_marcados),
                "epp_requerido": list(epp_marcados),
                "tarea_sif": str(_resolver_merge(hoja, fila, COL_SIF, mapa_merges) or "").strip().upper() == "SI",
            }
        )
        orden += 1
        fila += 1

    if not actividades:
        raise ErrorImportacionExcel(
            "No se encontró ninguna actividad en la tabla — revisa que el archivo tenga el formato esperado."
        )

    return {
        "planta_area": planta_area,
        "numero_pedido": numero_pedido,
        "gerente_proyecto": gerente_proyecto,
        "contacto_telefono": contacto_telefono,
        "fecha_elaboracion": fecha_elaboracion,
        "duracion_dias": duracion_dias,
        "descripcion_trabajo": descripcion_trabajo,
        "actividades": actividades,
        "avisos": avisos,
    }
