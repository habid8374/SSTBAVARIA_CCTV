import datetime
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from core.models import Empresa, PerfilUsuario

from .models import (
    ActividadMetodo,
    AutorizacionIngreso,
    DeclaracionMetodo,
    EmpresaContratista,
    FirmaMetodo,
    Funcionario,
    NotificacionInterna,
    RadicacionSeguridadSocial,
    RegistroAuditoria,
    Trabajador,
    nivel_riesgo,
)

Usuario = get_user_model()


class NivelRiesgoTests(TestCase):
    def test_muy_alto(self):
        clave, _ = nivel_riesgo(450)
        self.assertEqual(clave, "muy_alto")

    def test_alto(self):
        clave, _ = nivel_riesgo(300)
        self.assertEqual(clave, "alto")

    def test_considerable(self):
        clave, _ = nivel_riesgo(100)
        self.assertEqual(clave, "considerable")

    def test_posible(self):
        clave, _ = nivel_riesgo(50)
        self.assertEqual(clave, "posible")

    def test_bajo(self):
        clave, _ = nivel_riesgo(9)
        self.assertEqual(clave, "bajo")


class ActividadMetodoCalculoTests(TestCase):
    def test_riesgo_es_producto_p_f_i(self):
        actividad = ActividadMetodo(
            probabilidad_sin=6, frecuencia_sin=3, impacto_sin=3,
            probabilidad_con=1, frecuencia_con=3, impacto_con=1,
        )
        self.assertEqual(actividad.riesgo_sin, 54)
        self.assertEqual(actividad.riesgo_con, 3)


class RadicacionVencimientoTests(TestCase):
    def setUp(self):
        empresa = Empresa.objects.create(nombre="Bavaria Planta")
        contratista = EmpresaContratista.objects.create(empresa=empresa, nombre="SCEPSA")
        self.trabajador = Trabajador.objects.create(
            contratista=contratista, nombres="Ana", apellidos="Ríos", documento="123"
        )

    def test_vencida_true_si_la_fecha_ya_paso(self):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="ENERO",
            fecha_vencimiento=timezone.localdate() - datetime.timedelta(days=1),
        )
        self.assertTrue(radicacion.vencida)
        self.assertEqual(radicacion.dias_para_vencer, -1)

    def test_vencida_false_si_la_fecha_es_futura(self):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="ENERO",
            fecha_vencimiento=timezone.localdate() + datetime.timedelta(days=5),
        )
        self.assertFalse(radicacion.vencida)
        self.assertEqual(radicacion.dias_para_vencer, 5)

    def test_sin_fecha_de_vencimiento_no_esta_vencida_ni_tiene_dias(self):
        radicacion = RadicacionSeguridadSocial.objects.create(trabajador=self.trabajador, anio=2026, mes="ENERO")
        self.assertFalse(radicacion.vencida)
        self.assertIsNone(radicacion.dias_para_vencer)


class ApiTestsBase(TestCase):
    def setUp(self):
        # El throttle de login cuenta por IP y el test client siempre usa la
        # misma — sin esto, los _token() de tests anteriores se acumularían.
        cache.clear()
        self.empresa = Empresa.objects.create(nombre="Bavaria Planta")
        self.admin = Usuario.objects.create_superuser("admin", "admin@x.com", "clave12345")
        self.operador = Usuario.objects.create_user("operador1", "op@x.com", "clave12345")
        self.contratista = EmpresaContratista.objects.create(
            empresa=self.empresa, nombre="SCEPSA COLOMBIA SAS", nit="900588170-2"
        )
        self.trabajador = Trabajador.objects.create(
            contratista=self.contratista,
            nombres="Gerald Marcelo",
            apellidos="Garzón Beltrán",
            documento="80432071",
            eps="Nueva EPS",
            arl="Positiva",
            afp="Protección",
        )

    def _token(self, user):
        response = self.client.post(
            reverse("core:login"),
            {"username": user.username, "password": "clave12345"},
            content_type="application/json",
        )
        return response.data["token"]

    def _auth(self, user):
        return {"HTTP_AUTHORIZATION": f"Token {self._token(user)}"}


class CatalogosTests(ApiTestsBase):
    def test_catalogos_requiere_autenticacion(self):
        response = self.client.get(reverse("contratistas:catalogos"))
        self.assertEqual(response.status_code, 401)

    def test_catalogos_devuelve_listas(self):
        response = self.client.get(reverse("contratistas:catalogos"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["cursos_safety_academy"]), 7)
        self.assertIn("Trabajos en Altura > 1.8 m", response.data["permisos_trabajo"])
        self.assertEqual(len(response.data["roles_firma"]), 5)


class IndicadoresTests(ApiTestsBase):
    def test_cuenta_vencidas_y_por_vencer(self):
        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="ENERO",
            fecha_vencimiento=timezone.localdate() - datetime.timedelta(days=1),
        )
        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="FEBRERO",
            fecha_vencimiento=timezone.localdate() + datetime.timedelta(days=5),
        )
        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="MARZO",
            fecha_vencimiento=timezone.localdate() + datetime.timedelta(days=60),
        )

        response = self.client.get(reverse("contratistas:indicadores"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["radicaciones_vencidas"], 1)
        self.assertEqual(response.data["radicaciones_por_vencer"], 1)

    def test_excluye_rechazadas(self):
        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="ENERO",
            fecha_vencimiento=timezone.localdate() - datetime.timedelta(days=1),
            estado="rechazada",
        )
        response = self.client.get(reverse("contratistas:indicadores"), **self._auth(self.operador))
        self.assertEqual(response.data["radicaciones_vencidas"], 0)

    def test_requiere_autenticacion(self):
        response = self.client.get(reverse("contratistas:indicadores"))
        self.assertEqual(response.status_code, 401)

    def test_cuenta_certificaciones_de_trabajadores_vencidas_y_por_vencer(self):
        self.trabajador.fecha_vencimiento_examen_medico = timezone.localdate() - datetime.timedelta(days=1)
        self.trabajador.fecha_vencimiento_certificacion_alturas = timezone.localdate() + datetime.timedelta(days=5)
        self.trabajador.save(
            update_fields=["fecha_vencimiento_examen_medico", "fecha_vencimiento_certificacion_alturas"]
        )
        response = self.client.get(reverse("contratistas:indicadores"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["examenes_medicos_vencidos"], 1)
        self.assertEqual(response.data["certificaciones_alturas_por_vencer"], 1)

    def test_no_cuenta_certificaciones_de_trabajadores_inactivos(self):
        self.trabajador.activo = False
        self.trabajador.fecha_vencimiento_examen_medico = timezone.localdate() - datetime.timedelta(days=1)
        self.trabajador.save(update_fields=["activo", "fecha_vencimiento_examen_medico"])
        response = self.client.get(reverse("contratistas:indicadores"), **self._auth(self.operador))
        self.assertEqual(response.data["examenes_medicos_vencidos"], 0)


class IndicadoresDashboardTests(ApiTestsBase):
    def test_requiere_autenticacion(self):
        response = self.client.get(reverse("contratistas:indicadores_dashboard"))
        self.assertEqual(response.status_code, 401)

    def test_cuenta_trabajadores_con_cursos_obligatorios_pendientes(self):
        from .models import CursoSafetyAcademy

        CursoSafetyAcademy.objects.filter(clave="induccion_sst").update(obligatorio=True)
        # self.trabajador (de ApiTestsBase) no tiene ningún curso completado.
        response = self.client.get(reverse("contratistas:indicadores_dashboard"), **self._auth(self.operador))
        self.assertEqual(response.data["trabajadores_con_cursos_pendientes"], 1)

    def test_estructura_basica_sin_datos(self):
        response = self.client.get(reverse("contratistas:indicadores_dashboard"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["contratistas_activos"], 1)
        self.assertEqual(response.data["trabajadores_activos"], 1)
        self.assertEqual(response.data["riesgo_promedio_sin"], 0)
        self.assertIsNone(response.data["tiempo_promedio_aprobacion_dias"])
        self.assertEqual(len(response.data["tendencia_mensual"]), 6)
        self.assertEqual(len(response.data["top_riesgos"]), 0)

    def test_riesgo_promedio_y_top_riesgos(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Riesgo alto",
            probabilidad_sin=10, frecuencia_sin=10, impacto_sin=15,
        )
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=1, secuencia="Riesgo bajo",
            probabilidad_sin=1, frecuencia_sin=1, impacto_sin=1,
        )
        response = self.client.get(reverse("contratistas:indicadores_dashboard"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        # (1500 + 1) / 2 = 750.5
        self.assertEqual(response.data["riesgo_promedio_sin"], 750.5)
        self.assertEqual(response.data["top_riesgos"][0]["secuencia"], "Riesgo alto")
        self.assertEqual(response.data["top_riesgos"][0]["riesgo_sin"], 1500)

    def test_por_contratista_cuenta_pendientes(self):
        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO", estado="pendiente"
        )
        DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
            estado="enviada",
        )
        response = self.client.get(reverse("contratistas:indicadores_dashboard"), **self._auth(self.operador))
        fila = response.data["por_contratista"][0]
        self.assertEqual(fila["contratista"], "SCEPSA COLOMBIA SAS")
        self.assertEqual(fila["trabajadores"], 1)
        self.assertEqual(fila["radicaciones_pendientes"], 1)
        self.assertEqual(fila["declaraciones_pendientes"], 1)

    def test_tendencia_mensual_cuenta_mes_actual(self):
        DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=timezone.localdate(),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.get(reverse("contratistas:indicadores_dashboard"), **self._auth(self.operador))
        self.assertEqual(response.data["tendencia_mensual"][-1]["declaraciones"], 1)


class ReglasConfigurablesTests(ApiTestsBase):
    def test_catalogos_solo_incluye_cursos_y_permisos_activos(self):
        from .models import ConfiguracionAlertas, CursoSafetyAcademy, EquipoProteccionPersonal, PermisoTrabajo

        CursoSafetyAcademy.objects.create(clave="curso_inactivo", etiqueta="Curso inactivo", activo=False)
        PermisoTrabajo.objects.create(nombre="Permiso inactivo", activo=False)
        EquipoProteccionPersonal.objects.create(nombre="EPP inactivo", activo=False)
        response = self.client.get(reverse("contratistas:catalogos"), **self._auth(self.operador))
        claves = [c["clave"] for c in response.data["cursos_safety_academy"]]
        self.assertNotIn("curso_inactivo", claves)
        self.assertNotIn("Permiso inactivo", response.data["permisos_trabajo"])
        self.assertNotIn("EPP inactivo", response.data["equipos_epp"])

    def test_crear_curso(self):
        response = self.client.post(
            reverse("contratistas:cursos_lista"),
            {"clave": "nuevo_curso", "etiqueta": "Nuevo curso", "orden": 10},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)

    def test_operador_no_puede_eliminar_curso(self):
        from .models import CursoSafetyAcademy

        curso = CursoSafetyAcademy.objects.create(clave="x", etiqueta="X")
        response = self.client.delete(
            reverse("contratistas:cursos_detalle", args=[curso.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_puede_eliminar_permiso(self):
        from .models import PermisoTrabajo

        permiso = PermisoTrabajo.objects.create(nombre="X")
        response = self.client.delete(
            reverse("contratistas:permisos_detalle", args=[permiso.pk]), **self._auth(self.admin)
        )
        self.assertEqual(response.status_code, 204)

    def test_crear_epp(self):
        response = self.client.post(
            reverse("contratistas:equipos_epp_lista"),
            {"nombre": "Nuevo EPP de prueba", "orden": 1},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)

    def test_operador_no_puede_eliminar_epp(self):
        from .models import EquipoProteccionPersonal

        epp = EquipoProteccionPersonal.objects.create(nombre="Guantes")
        response = self.client.delete(
            reverse("contratistas:equipos_epp_detalle", args=[epp.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_puede_eliminar_epp(self):
        from .models import EquipoProteccionPersonal

        epp = EquipoProteccionPersonal.objects.create(nombre="Guantes")
        response = self.client.delete(
            reverse("contratistas:equipos_epp_detalle", args=[epp.pk]), **self._auth(self.admin)
        )
        self.assertEqual(response.status_code, 204)

    def test_configuracion_alertas_operador_solo_lee(self):
        response = self.client.get(reverse("contratistas:configuracion_alertas"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["dias_alerta_vencimiento"], 15)

        response = self.client.patch(
            reverse("contratistas:configuracion_alertas"),
            {"dias_alerta_vencimiento": 30},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 403)

    def test_configuracion_alertas_admin_puede_editar_y_afecta_indicadores(self):
        response = self.client.patch(
            reverse("contratistas:configuracion_alertas"),
            {"dias_alerta_vencimiento": 30},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)

        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="ENERO",
            fecha_vencimiento=timezone.localdate() + datetime.timedelta(days=25),
        )
        response = self.client.get(reverse("contratistas:indicadores"), **self._auth(self.operador))
        # con el umbral por defecto (15 días) esta planilla NO contaría como "por vencer";
        # con el nuevo umbral configurado (30 días) sí debe contar.
        self.assertEqual(response.data["radicaciones_por_vencer"], 1)


class FuncionarioTests(ApiTestsBase):
    def test_crear_funcionario(self):
        response = self.client.post(
            reverse("contratistas:funcionarios_lista"),
            {"nombre": "Jenny Quintero", "cargo": "Coordinadora HSEQ", "rol_firma": "delegado_abi", "correo": "jq@x.com"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(Funcionario.objects.count(), 1)

    def test_listar_filtra_por_rol_y_activo(self):
        Funcionario.objects.create(
            empresa=self.empresa, nombre="Jenny", rol_firma="delegado_abi", activo=True
        )
        Funcionario.objects.create(
            empresa=self.empresa, nombre="Carlos", rol_firma="lider_area", activo=False
        )
        response = self.client.get(
            reverse("contratistas:funcionarios_lista") + "?rol_firma=delegado_abi",
            **self._auth(self.operador),
        )
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["nombre"], "Jenny")

        response = self.client.get(
            reverse("contratistas:funcionarios_lista") + "?activo=true", **self._auth(self.operador)
        )
        self.assertEqual(len(response.data), 1)

    def test_operador_no_puede_eliminar(self):
        funcionario = Funcionario.objects.create(empresa=self.empresa, nombre="Jenny", rol_firma="delegado_abi")
        response = self.client.delete(
            reverse("contratistas:funcionarios_detalle", args=[funcionario.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_puede_eliminar(self):
        funcionario = Funcionario.objects.create(empresa=self.empresa, nombre="Jenny", rol_firma="delegado_abi")
        response = self.client.delete(
            reverse("contratistas:funcionarios_detalle", args=[funcionario.pk]), **self._auth(self.admin)
        )
        self.assertEqual(response.status_code, 204)

    def test_requiere_autenticacion(self):
        response = self.client.get(reverse("contratistas:funcionarios_lista"))
        self.assertEqual(response.status_code, 401)


class NotificacionPendienteTests(ApiTestsBase):
    @override_settings(BREVO_API_KEY="clave-de-prueba")
    @patch("camaras_ia.notificaciones.urllib.request.urlopen")
    def test_radicar_avisa_al_correo_revisor_configurado(self, mock_urlopen):
        from .models import ConfiguracionAlertas

        mock_urlopen.return_value.__enter__.return_value.status = 201
        ConfiguracionAlertas.obtener()
        ConfiguracionAlertas.objects.update(correo_revisor="revisor@empresa.com")

        response = self.client.post(
            reverse("contratistas:radicaciones_lista"),
            {"trabajador": self.trabajador.pk, "anio": 2026, "mes": "AGOSTO"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        mock_urlopen.assert_called_once()

    @patch("camaras_ia.notificaciones.urllib.request.urlopen")
    def test_radicar_sin_correo_revisor_configurado_no_avisa(self, mock_urlopen):
        response = self.client.post(
            reverse("contratistas:radicaciones_lista"),
            {"trabajador": self.trabajador.pk, "anio": 2026, "mes": "AGOSTO"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        mock_urlopen.assert_not_called()

    @override_settings(BREVO_API_KEY="clave-de-prueba")
    @patch("camaras_ia.notificaciones.urllib.request.urlopen")
    def test_enviar_declaracion_avisa_al_correo_revisor(self, mock_urlopen):
        from .models import ConfiguracionAlertas

        mock_urlopen.return_value.__enter__.return_value.status = 201
        ConfiguracionAlertas.obtener()
        ConfiguracionAlertas.objects.update(correo_revisor="revisor@empresa.com")

        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "enviada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)
        mock_urlopen.assert_called_once()

    @patch("contratistas.views.notificar_declaracion_pendiente")
    def test_aprobar_no_dispara_aviso_de_pendiente(self, mock_notificar_pendiente):
        from .models import FirmaMetodo

        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)
        mock_notificar_pendiente.assert_not_called()


class NotificacionInternaTests(ApiTestsBase):
    """La bandeja propia del dashboard no depende de que correo_revisor esté
    configurado ni de que Brevo funcione — es un segundo canal, siempre
    activo, para que el personal interno vea qué le falta por revisar."""

    def test_radicar_crea_notificacion_pendiente_aunque_no_haya_correo_revisor(self):
        response = self.client.post(
            reverse("contratistas:radicaciones_lista"),
            {"trabajador": self.trabajador.pk, "anio": 2026, "mes": "AGOSTO"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        notificacion = NotificacionInterna.objects.latest("creada_en")
        self.assertEqual(notificacion.tipo, NotificacionInterna.Tipo.RADICACION_PENDIENTE)
        self.assertFalse(notificacion.leida)

    def test_enviar_declaracion_nueva_crea_notificacion_pendiente(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "enviada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)
        notificacion = NotificacionInterna.objects.latest("creada_en")
        self.assertEqual(notificacion.tipo, NotificacionInterna.Tipo.DECLARACION_PENDIENTE)
        self.assertEqual(notificacion.objeto_id, declaracion.pk)

    def test_reenviar_declaracion_rechazada_crea_notificacion_de_subsanacion(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
            estado=DeclaracionMetodo.Estado.RECHAZADA,
            observaciones="Falta el permiso de trabajo en altura.",
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "enviada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)
        notificacion = NotificacionInterna.objects.latest("creada_en")
        self.assertEqual(notificacion.tipo, NotificacionInterna.Tipo.DECLARACION_SUBSANADA)

    def test_aprobar_no_crea_notificacion_de_pendiente(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )
        self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertFalse(
            NotificacionInterna.objects.filter(
                tipo__in=[NotificacionInterna.Tipo.DECLARACION_PENDIENTE, NotificacionInterna.Tipo.DECLARACION_SUBSANADA]
            ).exists()
        )

    def test_lista_requiere_personal_interno(self):
        portal_user = Usuario.objects.create_user("portal", "portal@x.com", "clave12345")
        from core.models import PerfilUsuario

        portal_user.perfil.rol = PerfilUsuario.Rol.CONTRATISTA
        portal_user.perfil.contratista = self.contratista
        portal_user.perfil.save(update_fields=["rol", "contratista"])

        response = self.client.get(
            reverse("contratistas:notificaciones_internas_lista"), **self._auth(portal_user)
        )
        self.assertEqual(response.status_code, 403)

    def test_marcar_notificacion_leida(self):
        notificacion = NotificacionInterna.objects.create(
            tipo=NotificacionInterna.Tipo.RADICACION_PENDIENTE,
            mensaje="Algo pendiente",
            modelo="RadicacionSeguridadSocial",
            objeto_id=1,
        )
        response = self.client.post(
            reverse("contratistas:notificaciones_internas_marcar_leida", args=[notificacion.pk]),
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200, response.data)
        notificacion.refresh_from_db()
        self.assertTrue(notificacion.leida)

    def test_marcar_todas_leidas(self):
        NotificacionInterna.objects.create(
            tipo=NotificacionInterna.Tipo.RADICACION_PENDIENTE, mensaje="A", modelo="X", objeto_id=1
        )
        NotificacionInterna.objects.create(
            tipo=NotificacionInterna.Tipo.DECLARACION_PENDIENTE, mensaje="B", modelo="Y", objeto_id=2
        )
        response = self.client.post(
            reverse("contratistas:notificaciones_internas_marcar_todas_leidas"), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 204)
        self.assertEqual(NotificacionInterna.objects.filter(leida=False).count(), 0)


class EmpresaContratistaTests(ApiTestsBase):
    def test_lista(self):
        response = self.client.get(reverse("contratistas:empresas_lista"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["trabajadores_count"], 1)

    def test_crear(self):
        response = self.client.post(
            reverse("contratistas:empresas_lista"),
            {"nombre": "Gestión y Control Integral del Riesgo SAS", "nit": "900123456-1"},
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(EmpresaContratista.objects.count(), 2)

    def test_operador_no_puede_eliminar_empresa(self):
        url = reverse("contratistas:empresas_detalle", args=[self.contratista.pk])
        response = self.client.delete(url, **self._auth(self.operador))
        self.assertEqual(response.status_code, 403)
        self.assertTrue(EmpresaContratista.objects.filter(pk=self.contratista.pk).exists())

    def test_admin_puede_eliminar_empresa(self):
        url = reverse("contratistas:empresas_detalle", args=[self.contratista.pk])
        response = self.client.delete(url, **self._auth(self.admin))
        self.assertEqual(response.status_code, 204)
        self.assertFalse(EmpresaContratista.objects.filter(pk=self.contratista.pk).exists())


class RegistroAuditoriaTests(ApiTestsBase):
    def test_crear_empresa_queda_registrado(self):
        self.client.post(
            reverse("contratistas:empresas_lista"),
            {"nombre": "Gestión y Control Integral del Riesgo SAS", "nit": "900123456-1"},
            **self._auth(self.operador),
        )
        registro = RegistroAuditoria.objects.filter(modelo="EmpresaContratista", accion="creado").latest("fecha")
        self.assertEqual(registro.objeto_str, "Gestión y Control Integral del Riesgo SAS")
        self.assertEqual(registro.usuario, self.operador)

    def test_editar_trabajador_registra_solo_los_campos_que_cambiaron(self):
        self.client.patch(
            reverse("contratistas:trabajadores_detalle", args=[self.trabajador.pk]),
            {"eps": "Sura"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        registro = RegistroAuditoria.objects.filter(modelo="Trabajador", accion="actualizado").latest("fecha")
        self.assertEqual(set(registro.cambios.keys()), {"eps"})
        self.assertEqual(registro.cambios["eps"], {"antes": "Nueva EPS", "despues": "Sura"})

    def test_editar_sin_cambios_reales_no_crea_registro(self):
        cantidad_antes = RegistroAuditoria.objects.count()
        self.client.patch(
            reverse("contratistas:trabajadores_detalle", args=[self.trabajador.pk]),
            {"eps": self.trabajador.eps},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(RegistroAuditoria.objects.count(), cantidad_antes)

    def test_eliminar_empresa_queda_registrado_aunque_ya_no_exista(self):
        pk = self.contratista.pk
        nombre = self.contratista.nombre
        self.client.delete(reverse("contratistas:empresas_detalle", args=[pk]), **self._auth(self.admin))
        registro = RegistroAuditoria.objects.filter(modelo="EmpresaContratista", accion="eliminado").latest("fecha")
        self.assertEqual(registro.objeto_id, pk)
        self.assertEqual(registro.objeto_str, nombre)

    def test_aprobar_radicacion_registra_cambio_de_estado(self):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="ENERO", estado=RadicacionSeguridadSocial.Estado.PENDIENTE
        )
        self.client.post(
            reverse("contratistas:radicaciones_aprobar", args=[radicacion.pk]),
            {},
            content_type="application/json",
            **self._auth(self.operador),
        )
        registro = RegistroAuditoria.objects.filter(
            modelo="RadicacionSeguridadSocial", objeto_id=radicacion.pk, accion="actualizado"
        ).latest("fecha")
        self.assertEqual(registro.cambios["estado"], {"antes": "pendiente", "despues": "aprobada"})

    def test_operador_no_puede_ver_auditoria(self):
        response = self.client.get(reverse("contratistas:auditoria_lista"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 403)

    def test_admin_puede_ver_y_filtrar_auditoria(self):
        self.client.post(
            reverse("contratistas:empresas_lista"),
            {"nombre": "Otra Contratista SAS"},
            **self._auth(self.admin),
        )
        response = self.client.get(
            reverse("contratistas:auditoria_lista") + "?modelo=EmpresaContratista",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(all(r["modelo"] == "EmpresaContratista" for r in response.data))


class TrabajadorTests(ApiTestsBase):
    def test_cursos_pendientes_lista_obligatorios_incompletos(self):
        from .models import CursoSafetyAcademy

        CursoSafetyAcademy.objects.filter(clave__in=["induccion_sst", "epp"]).update(obligatorio=True)
        self.trabajador.cursos_safety_academy = {"induccion_sst": "2026-01-01", "epp": None}
        self.trabajador.save()

        response = self.client.get(
            reverse("contratistas:trabajadores_detalle", args=[self.trabajador.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 200)
        claves = [c["clave"] for c in response.data["cursos_pendientes"]]
        self.assertEqual(claves, ["epp"])

    def test_curso_no_obligatorio_no_cuenta_como_pendiente(self):
        from .models import CursoSafetyAcademy

        CursoSafetyAcademy.objects.filter(clave="epp").update(obligatorio=False)
        response = self.client.get(
            reverse("contratistas:trabajadores_detalle", args=[self.trabajador.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.data["cursos_pendientes"], [])

    def test_examen_medico_vencido_expuesto_en_el_detalle(self):
        self.trabajador.fecha_vencimiento_examen_medico = timezone.localdate() - datetime.timedelta(days=1)
        self.trabajador.save(update_fields=["fecha_vencimiento_examen_medico"])
        response = self.client.get(
            reverse("contratistas:trabajadores_detalle", args=[self.trabajador.pk]), **self._auth(self.operador)
        )
        self.assertTrue(response.data["examen_medico_vencido"])
        self.assertEqual(response.data["dias_para_vencer_examen_medico"], -1)

    def test_certificacion_alturas_vigente_expuesta_en_el_detalle(self):
        self.trabajador.fecha_vencimiento_certificacion_alturas = timezone.localdate() + datetime.timedelta(days=10)
        self.trabajador.save(update_fields=["fecha_vencimiento_certificacion_alturas"])
        response = self.client.get(
            reverse("contratistas:trabajadores_detalle", args=[self.trabajador.pk]), **self._auth(self.operador)
        )
        self.assertFalse(response.data["certificacion_alturas_vencida"])
        self.assertEqual(response.data["dias_para_vencer_certificacion_alturas"], 10)

    def test_sin_fecha_de_vencimiento_no_marca_vencido(self):
        response = self.client.get(
            reverse("contratistas:trabajadores_detalle", args=[self.trabajador.pk]), **self._auth(self.operador)
        )
        self.assertFalse(response.data["examen_medico_vencido"])
        self.assertIsNone(response.data["dias_para_vencer_examen_medico"])

    def test_crear_trabajador(self):
        response = self.client.post(
            reverse("contratistas:trabajadores_lista"),
            {
                "contratista": self.contratista.pk,
                "nombres": "Luis Alfonso",
                "apellidos": "Estepa Patiño",
                "documento": "80431911",
                "cursos_safety_academy": {"induccion_sst": "2026-08-01"},
                "autorizacion_datos": True,
            },
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(Trabajador.objects.count(), 2)
        trabajador = Trabajador.objects.get(documento="80431911")
        self.assertTrue(trabajador.autorizacion_datos)
        self.assertIsNotNone(trabajador.autorizacion_datos_en)

    def test_crear_trabajador_sin_autorizacion_devuelve_400(self):
        response = self.client.post(
            reverse("contratistas:trabajadores_lista"),
            {
                "contratista": self.contratista.pk,
                "nombres": "Luis Alfonso",
                "apellidos": "Estepa Patiño",
                "documento": "80431911",
            },
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("autorizacion_datos", response.data)
        self.assertFalse(Trabajador.objects.filter(documento="80431911").exists())

    def test_editar_no_reexige_autorizacion_y_no_recalcula_fecha(self):
        fecha_original = datetime.datetime(2026, 1, 1, tzinfo=datetime.timezone.utc)
        trabajador = Trabajador.objects.create(
            contratista=self.contratista,
            nombres="Ana",
            apellidos="Ríos",
            documento="999",
            autorizacion_datos=True,
            autorizacion_datos_en=fecha_original,
        )

        response = self.client.patch(
            reverse("contratistas:trabajadores_detalle", args=[trabajador.pk]),
            {"eps": "Sura"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200, response.data)
        trabajador.refresh_from_db()
        self.assertEqual(trabajador.autorizacion_datos_en, fecha_original)

    def test_crear_trabajador_con_evidencia_de_autorizacion(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        evidencia = SimpleUploadedFile("autorizacion.pdf", b"%PDF-1.4 contenido falso", content_type="application/pdf")
        response = self.client.post(
            reverse("contratistas:trabajadores_lista"),
            {
                "contratista": self.contratista.pk,
                "nombres": "Luis Alfonso",
                "apellidos": "Estepa Patiño",
                "documento": "80431911",
                "autorizacion_datos": True,
                "soporte_autorizacion_datos": evidencia,
            },
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        trabajador = Trabajador.objects.get(documento="80431911")
        self.assertTrue(trabajador.soporte_autorizacion_datos.name)

    def test_evidencia_de_autorizacion_con_extension_invalida_devuelve_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        archivo = SimpleUploadedFile("malware.exe", b"MZ contenido falso", content_type="application/octet-stream")
        response = self.client.post(
            reverse("contratistas:trabajadores_lista"),
            {
                "contratista": self.contratista.pk,
                "nombres": "Luis Alfonso",
                "apellidos": "Estepa Patiño",
                "documento": "80431911",
                "autorizacion_datos": True,
                "soporte_autorizacion_datos": archivo,
            },
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("soporte_autorizacion_datos", response.data)

    def test_filtro_por_contratista(self):
        otro_contratista = EmpresaContratista.objects.create(empresa=self.empresa, nombre="Otra SAS")
        Trabajador.objects.create(
            contratista=otro_contratista, nombres="X", apellidos="Y", documento="1"
        )
        response = self.client.get(
            reverse("contratistas:trabajadores_lista"),
            {"contratista": self.contratista.pk},
            **self._auth(self.operador),
        )
        self.assertEqual(len(response.data), 1)


class RadicacionSeguridadSocialTests(ApiTestsBase):
    def test_radicar_y_aprobar(self):
        response = self.client.post(
            reverse("contratistas:radicaciones_lista"),
            {
                "trabajador": self.trabajador.pk,
                "anio": 2026,
                "mes": "AGOSTO",
                "numero_planilla": "94075251",
                "fecha_vencimiento": "2026-09-19",
            },
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201)
        radicacion_id = response.data["id"]
        self.assertEqual(response.data["estado"], "pendiente")

        response = self.client.post(
            reverse("contratistas:radicaciones_aprobar", args=[radicacion_id]),
            {"observaciones": "Todo en orden"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["estado"], "aprobada")
        radicacion = RadicacionSeguridadSocial.objects.get(pk=radicacion_id)
        self.assertIsNotNone(radicacion.revisada_en)

    def test_rechazar(self):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO"
        )
        response = self.client.post(
            reverse("contratistas:radicaciones_rechazar", args=[radicacion.pk]),
            {"observaciones": "Planilla vencida"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["estado"], "rechazada")

    def test_rechazar_sin_observaciones_devuelve_400(self):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO"
        )
        response = self.client.post(
            reverse("contratistas:radicaciones_rechazar", args=[radicacion.pk]),
            {},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)
        radicacion.refresh_from_db()
        self.assertEqual(radicacion.estado, "pendiente")

    def test_aprobar_no_exige_observaciones(self):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO"
        )
        response = self.client.post(
            reverse("contratistas:radicaciones_aprobar", args=[radicacion.pk]),
            {},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200, response.data)

    @override_settings(BREVO_API_KEY="clave-de-prueba")
    @patch("camaras_ia.notificaciones.urllib.request.urlopen")
    def test_aprobar_notifica_al_correo_de_contacto(self, mock_urlopen):
        mock_urlopen.return_value.__enter__.return_value.status = 201
        self.contratista.contacto_correo = "contacto@contratista.com"
        self.contratista.save()
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO"
        )
        response = self.client.post(
            reverse("contratistas:radicaciones_aprobar", args=[radicacion.pk]),
            {"observaciones": "Todo en orden"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200)
        mock_urlopen.assert_called_once()

    @patch("camaras_ia.notificaciones.urllib.request.urlopen")
    def test_sin_correo_de_contacto_no_intenta_notificar(self, mock_urlopen):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO"
        )
        response = self.client.post(
            reverse("contratistas:radicaciones_aprobar", args=[radicacion.pk]),
            {"observaciones": "Todo en orden"},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200)
        mock_urlopen.assert_not_called()

    def test_filtra_por_vencida(self):
        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="ENERO",
            fecha_vencimiento=timezone.localdate() - datetime.timedelta(days=1),
        )
        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador,
            anio=2026,
            mes="FEBRERO",
            fecha_vencimiento=timezone.localdate() + datetime.timedelta(days=10),
        )

        response = self.client.get(
            reverse("contratistas:radicaciones_lista"), {"vencida": "true"}, **self._auth(self.operador)
        )
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["mes"], "ENERO")

    def test_exportar_devuelve_xlsx_con_las_filas(self):
        import openpyxl
        from io import BytesIO

        RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO", numero_planilla="94075251"
        )
        response = self.client.get(reverse("contratistas:radicaciones_exportar"), **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        libro = openpyxl.load_workbook(BytesIO(response.content))
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        self.assertEqual(filas[0][0], "Contratista")  # encabezado
        self.assertIn("94075251", filas[1])

    def test_exportar_respeta_filtro_de_contratista(self):
        from io import BytesIO

        import openpyxl

        otro_contratista = EmpresaContratista.objects.create(empresa=self.empresa, nombre="Otra SAS")
        otro_trabajador = Trabajador.objects.create(
            contratista=otro_contratista, nombres="X", apellidos="Y", documento="999"
        )
        RadicacionSeguridadSocial.objects.create(trabajador=self.trabajador, anio=2026, mes="AGOSTO")
        RadicacionSeguridadSocial.objects.create(trabajador=otro_trabajador, anio=2026, mes="AGOSTO")

        response = self.client.get(
            reverse("contratistas:radicaciones_exportar"),
            {"contratista": self.contratista.pk},
            **self._auth(self.operador),
        )
        libro = openpyxl.load_workbook(BytesIO(response.content))
        filas = list(libro.active.iter_rows(values_only=True))
        self.assertEqual(len(filas), 2)  # encabezado + 1 fila

    def test_requiere_autenticacion(self):
        response = self.client.get(reverse("contratistas:radicaciones_lista"))
        self.assertEqual(response.status_code, 401)

    def test_rechaza_extension_de_archivo_no_permitida(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        archivo = SimpleUploadedFile("malware.exe", b"MZ contenido falso", content_type="application/octet-stream")
        response = self.client.post(
            reverse("contratistas:radicaciones_lista"),
            {"trabajador": self.trabajador.pk, "anio": 2026, "mes": "AGOSTO", "soporte_pago": archivo},
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("soporte_pago", response.data)

    def test_solo_admin_puede_eliminar_radicacion(self):
        radicacion = RadicacionSeguridadSocial.objects.create(
            trabajador=self.trabajador, anio=2026, mes="AGOSTO"
        )
        url = reverse("contratistas:radicaciones_detalle", args=[radicacion.pk])

        respuesta_operador = self.client.delete(url, **self._auth(self.operador))
        self.assertEqual(respuesta_operador.status_code, 403)

        respuesta_admin = self.client.delete(url, **self._auth(self.admin))
        self.assertEqual(respuesta_admin.status_code, 204)


class AlertasAutomaticasTests(ApiTestsBase):
    """Motor de alertas automáticas — nunca decide por sí solo, solo genera
    advertencias informativas con un motivo sugerido. Ver
    contratistas/alertas_automaticas.py."""

    def _declaracion(self):
        return DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )

    def test_altura_sin_epp_caida_genera_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Subir al techo",
            permisos_requeridos=["Trabajos en Altura > 1.8 m"],
            epp_requerido=["Casco de seguridad"],
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("altura_sin_epp_caida", [a["codigo"] for a in alertas])

    def test_altura_con_epp_caida_no_genera_esa_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Subir al techo",
            permisos_requeridos=["Trabajos en Altura > 1.8 m"],
            epp_requerido=["Otros: Equipo contra caídas (Arnés de seguridad, línea retráctil, doble gancho)"],
        )
        alertas = generar_alertas(declaracion)
        self.assertNotIn("altura_sin_epp_caida", [a["codigo"] for a in alertas])

    def test_excavacion_sin_medidas_genera_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Zanja para tubería",
            permisos_requeridos=["Excavaciones o Demolición"],
            medidas_mitigacion="",
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("excavacion_sin_medidas", [a["codigo"] for a in alertas])

    def test_riesgo_alto_con_mitigacion_genera_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Trabajo riesgoso",
            probabilidad_con=10,
            frecuencia_con=6,
            impacto_con=15,
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("riesgo_alto_con_mitigacion", [a["codigo"] for a in alertas])

    def test_riesgo_bajo_con_mitigacion_no_genera_esa_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Trabajo controlado",
            probabilidad_con=1,
            frecuencia_con=1,
            impacto_con=1,
        )
        alertas = generar_alertas(declaracion)
        self.assertNotIn("riesgo_alto_con_mitigacion", [a["codigo"] for a in alertas])

    def test_sif_sin_firma_seguridad_genera_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(declaracion=declaracion, orden=0, secuencia="Trabajo SIF", tarea_sif=True)
        alertas = generar_alertas(declaracion)
        self.assertIn("sif_sin_firma_seguridad", [a["codigo"] for a in alertas])

    def test_sif_con_firma_seguridad_vigente_no_genera_esa_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(declaracion=declaracion, orden=0, secuencia="Trabajo SIF", tarea_sif=True)
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="seguridad_planta", nombre_firmante="Ana", firmante_usuario=self.admin
        )
        alertas = generar_alertas(declaracion)
        self.assertNotIn("sif_sin_firma_seguridad", [a["codigo"] for a in alertas])

    def test_texto_sugiere_altura_sin_permiso_genera_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Trabajo en el techo de la nave"
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("texto_sugiere_altura_sin_permiso", [a["codigo"] for a in alertas])

    def test_altura_sobre_1_8m_sin_permiso_genera_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Subir al techo", altura_trabajo_metros=2.5
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("altura_sobre_1_8m_sin_permiso", [a["codigo"] for a in alertas])

    def test_altura_bajo_1_8m_no_genera_esa_alerta(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Trabajo controlado", altura_trabajo_metros=1.2
        )
        alertas = generar_alertas(declaracion)
        self.assertNotIn("altura_sobre_1_8m_sin_permiso", [a["codigo"] for a in alertas])

    def test_altura_sobre_4m_requiere_zbs(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Trabajo en techo alto",
            altura_trabajo_metros=4.5,
            permisos_requeridos=["Trabajos en Altura > 1.8 m"],
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("altura_sobre_4m_requiere_zbs", [a["codigo"] for a in alertas])

    def test_altura_sin_diligenciar_no_genera_alertas_numericas(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(declaracion=declaracion, orden=0, secuencia="Trabajo en bodega")
        alertas = generar_alertas(declaracion)
        codigos = [a["codigo"] for a in alertas]
        self.assertNotIn("altura_sobre_1_8m_sin_permiso", codigos)
        self.assertNotIn("altura_sobre_4m_requiere_zbs", codigos)

    def test_excavacion_sobre_1_2m_requiere_salida_emergencia(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Zanja", profundidad_excavacion_metros=1.5
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("excavacion_sobre_1_2m_salida_emergencia", [a["codigo"] for a in alertas])

    def test_excavacion_sobre_1_3m_requiere_reten_exterior(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Zanja profunda", profundidad_excavacion_metros=2.0
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("excavacion_sobre_1_3m_reten_exterior", [a["codigo"] for a in alertas])

    def test_excavacion_sobre_5m_requiere_andamio(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Excavación profunda", profundidad_excavacion_metros=6.0
        )
        alertas = generar_alertas(declaracion)
        self.assertIn("excavacion_sobre_5m_requiere_andamio", [a["codigo"] for a in alertas])

    def test_excavacion_superficial_no_genera_alertas_numericas(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion, orden=0, secuencia="Zanja pequeña", profundidad_excavacion_metros=0.5
        )
        alertas = generar_alertas(declaracion)
        codigos = [a["codigo"] for a in alertas]
        self.assertNotIn("excavacion_sobre_1_2m_salida_emergencia", codigos)
        self.assertNotIn("excavacion_sobre_1_3m_reten_exterior", codigos)
        self.assertNotIn("excavacion_sobre_5m_requiere_andamio", codigos)

    def test_actividad_sin_problemas_no_genera_alertas(self):
        from .alertas_automaticas import generar_alertas

        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Limpieza de piso en bodega",
            probabilidad_con=1,
            frecuencia_con=1,
            impacto_con=1,
        )
        self.assertEqual(generar_alertas(declaracion), [])

    def test_endpoint_requiere_personal_interno(self):
        declaracion = self._declaracion()
        url = reverse("contratistas:declaraciones_alertas", args=[declaracion.pk])

        portal_user = Usuario.objects.create_user("portal_epp_test", "portal_epp@x.com", "clave12345")
        portal_user.perfil.rol = PerfilUsuario.Rol.CONTRATISTA
        portal_user.perfil.contratista = self.contratista
        portal_user.perfil.save(update_fields=["rol", "contratista"])

        response = self.client.get(url, **self._auth(portal_user))
        self.assertEqual(response.status_code, 403)

        response = self.client.get(url, **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)

    def test_endpoint_devuelve_alertas_de_la_declaracion(self):
        declaracion = self._declaracion()
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Subir al techo",
            permisos_requeridos=["Trabajos en Altura > 1.8 m"],
        )
        url = reverse("contratistas:declaraciones_alertas", args=[declaracion.pk])
        response = self.client.get(url, **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertIn("altura_sin_epp_caida", [a["codigo"] for a in response.data])


class DeclaracionMetodoTests(ApiTestsBase):
    def test_crear_con_actividades_anidadas(self):
        payload = {
            "contratista": self.contratista.pk,
            "planta_area": "Cervecería Tocancipá - Tapas",
            "fecha_elaboracion": "2026-07-11",
            "duracion_dias": 30,
            "descripcion_trabajo": "Instalación de polipasto / tecle",
            "actividades": [
                {
                    "orden": 0,
                    "secuencia": "1. Ingreso y salida a planta",
                    "probabilidad_sin": 6,
                    "frecuencia_sin": 3,
                    "impacto_sin": 3,
                    "medidas_mitigacion": "Uso de senderos peatonales",
                    "probabilidad_con": 3,
                    "frecuencia_con": 3,
                    "impacto_con": 1,
                    "permisos_requeridos": ["Trabajos en Altura > 1.8 m"],
                    "tarea_sif": True,
                },
                {
                    "orden": 1,
                    "secuencia": "2. Inducción por el cliente",
                    "probabilidad_sin": 6,
                    "frecuencia_sin": 2,
                    "impacto_sin": 3,
                },
            ],
        }
        response = self.client.post(
            reverse("contratistas:declaraciones_lista"),
            payload,
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        declaracion_id = response.data["id"]
        self.assertEqual(len(response.data["actividades"]), 2)
        self.assertEqual(response.data["actividades"][0]["riesgo_sin"], 54)
        self.assertEqual(response.data["actividades"][0]["nivel_riesgo_sin"]["clave"], "posible")

        declaracion = DeclaracionMetodo.objects.get(pk=declaracion_id)
        self.assertEqual(declaracion.actividades.count(), 2)

    def test_editar_reemplaza_actividades(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        ActividadMetodo.objects.create(declaracion=declaracion, orden=0, secuencia="Actividad vieja")

        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"actividades": [{"orden": 0, "secuencia": "Actividad nueva"}]},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200, response.data)
        declaracion.refresh_from_db()
        self.assertEqual(declaracion.actividades.count(), 1)
        self.assertEqual(declaracion.actividades.first().secuencia, "Actividad nueva")

    def test_aprobar_declaracion_sin_firmas_devuelve_400(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 400)
        declaracion.refresh_from_db()
        self.assertEqual(declaracion.estado, "borrador")

    def test_aprobar_declaracion_con_firma_permite(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["estado"], "aprobada")

    @override_settings(BREVO_API_KEY="clave-de-prueba")
    @patch("camaras_ia.notificaciones.urllib.request.urlopen")
    def test_aprobar_declaracion_notifica_al_contratista(self, mock_urlopen):
        mock_urlopen.return_value.__enter__.return_value.status = 201
        self.contratista.contacto_correo = "contacto@contratista.com"
        self.contratista.save()
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )

        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)
        mock_urlopen.assert_called_once()

    @patch("camaras_ia.notificaciones.urllib.request.urlopen")
    def test_sin_correo_de_contacto_no_intenta_notificar(self, mock_urlopen):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,  # contacto_correo vacío por defecto
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )

        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)
        mock_urlopen.assert_not_called()

    def test_editar_sin_cambiar_estado_no_exige_firma(self):
        """Solo la transición A 'aprobada' exige firma — editar cualquier
        otro campo (o dejar el estado en borrador) sigue libre."""
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"planta_area": "Cervecería Tocancipá"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 200, response.data)

    def test_pdf_devuelve_documento(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Ingreso a planta",
            probabilidad_sin=6,
            frecuencia_sin=3,
            impacto_sin=3,
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )

        response = self.client.get(
            reverse("contratistas:declaraciones_pdf", args=[declaracion.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_pdf_requiere_autenticacion(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.get(reverse("contratistas:declaraciones_pdf", args=[declaracion.pk]))
        self.assertEqual(response.status_code, 401)

    def test_excel_devuelve_documento(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        ActividadMetodo.objects.create(
            declaracion=declaracion,
            orden=0,
            secuencia="Ingreso a planta",
            probabilidad_sin=6,
            frecuencia_sin=3,
            impacto_sin=3,
            permisos_requeridos=["Trabajos en Altura > 1.8 m"],
            epp_requerido=["Casco de seguridad"],
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )

        response = self.client.get(
            reverse("contratistas:declaraciones_excel", args=[declaracion.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(f'declaracion-metodo-{declaracion.pk}.xlsx', response["Content-Disposition"])

    def test_excel_requiere_autenticacion(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.get(reverse("contratistas:declaraciones_excel", args=[declaracion.pk]))
        self.assertEqual(response.status_code, 401)

    def test_excel_sin_archivo_original_usa_formato_propio(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        import io

        from openpyxl import load_workbook

        response = self.client.get(
            reverse("contratistas:declaraciones_excel", args=[declaracion.pk]), **self._auth(self.operador)
        )
        libro = load_workbook(io.BytesIO(response.content))
        self.assertIn("Control del Documento", libro.sheetnames)

    def test_excel_con_archivo_original_lo_reutiliza_y_agrega_decision(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
            estado=DeclaracionMetodo.Estado.RECHAZADA,
            observaciones="Falta el permiso de trabajo en altura marcado.",
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        declaracion.archivo_origen_excel.save(
            "original.xlsx",
            SimpleUploadedFile("original.xlsx", _construir_excel_declaracion_prueba()),
            save=True,
        )

        import io

        from openpyxl import load_workbook

        response = self.client.get(
            reverse("contratistas:declaraciones_excel", args=[declaracion.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 200)
        libro = load_workbook(io.BytesIO(response.content))

        # El libro descargado debe ser el original (no el reconstruido por
        # generar_excel_declaracion) con una hoja "Decisión SST" agregada.
        self.assertNotIn("Control del Documento", libro.sheetnames)
        self.assertIn("Decisión SST", libro.sheetnames)
        hoja_original = libro["Declaración de Método"]
        self.assertEqual(hoja_original["B3"].value, "Planta Prueba")

        hoja_decision = libro["Decisión SST"]
        contenido = " ".join(
            str(hoja_decision.cell(row=f, column=c).value or "")
            for f in range(1, hoja_decision.max_row + 1)
            for c in range(1, hoja_decision.max_column + 1)
        )
        self.assertIn("Rechazada", contenido)
        self.assertIn("Falta el permiso de trabajo en altura marcado.", contenido)

    def test_firmar_declaracion(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.post(
            reverse("contratistas:declaraciones_firmar", args=[declaracion.pk]),
            {"rol": "supervisor_contratista", "nombre_firmante": "Andres Felipe Lujan", "consiento_firma": True},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(FirmaMetodo.objects.count(), 1)
        firma = FirmaMetodo.objects.first()
        self.assertEqual(firma.firmante_usuario, self.operador)
        self.assertTrue(firma.hash_documento)
        self.assertFalse(response.data["documento_modificado_despues_de_firmar"])
        self.assertEqual(response.data["firmante_usuario_nombre"], "operador1")

        # Firmar de nuevo el mismo rol reemplaza en vez de duplicar.
        response = self.client.post(
            reverse("contratistas:declaraciones_firmar", args=[declaracion.pk]),
            {"rol": "supervisor_contratista", "nombre_firmante": "Otro Nombre", "consiento_firma": True},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(FirmaMetodo.objects.count(), 1)
        self.assertEqual(FirmaMetodo.objects.first().nombre_firmante, "Otro Nombre")

    def test_firmar_sin_consentimiento_devuelve_400(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.post(
            reverse("contratistas:declaraciones_firmar", args=[declaracion.pk]),
            {"rol": "supervisor_contratista", "nombre_firmante": "Andres Felipe Lujan", "consiento_firma": False},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("consiento_firma", response.data)
        self.assertEqual(FirmaMetodo.objects.count(), 0)

    def test_editar_declaracion_despues_de_firmar_marca_documento_modificado(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        self.client.post(
            reverse("contratistas:declaraciones_firmar", args=[declaracion.pk]),
            {"rol": "supervisor_contratista", "nombre_firmante": "Ana", "consiento_firma": True},
            content_type="application/json",
            **self._auth(self.operador),
        )
        firma = FirmaMetodo.objects.first()
        self.assertFalse(firma.documento_modificado_despues_de_firmar)

        self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"descripcion_trabajo": "Instalación de pórtico modificada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        firma.refresh_from_db()
        self.assertTrue(firma.documento_modificado_despues_de_firmar)

        # Aprobar ya no debería dejarse — la firma quedó desactualizada.
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.admin),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("estado", response.data)

    def test_operador_no_puede_eliminar_declaracion(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.delete(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]), **self._auth(self.operador)
        )
        self.assertEqual(response.status_code, 403)
        self.assertTrue(DeclaracionMetodo.objects.filter(pk=declaracion.pk).exists())

    def test_admin_puede_eliminar_declaracion(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.delete(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]), **self._auth(self.admin)
        )
        self.assertEqual(response.status_code, 204)
        self.assertFalse(DeclaracionMetodo.objects.filter(pk=declaracion.pk).exists())
        registro = RegistroAuditoria.objects.filter(modelo="DeclaracionMetodo", accion="eliminado").latest("fecha")
        self.assertEqual(registro.usuario, self.admin)


def _construir_excel_declaracion_prueba():
    """Arma en memoria un Excel mínimo con el mismo formato real del
    cliente (celdas del encabezado combinadas con texto tipo
    "ETIQUETA: valor", tabla de actividades desde la fila 14, hoja de
    Firmas/Permisos/EPP con columnas etiqueta/marca) — lo justo para que
    contratistas.importar_declaracion_excel.parsear_excel_declaracion lo
    reconozca, sin depender de los archivos reales del cliente."""
    import io

    from openpyxl import Workbook

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Declaración de Método"
    hoja["B3"] = "Planta Prueba"
    hoja["A6"] = (
        "NOMBRE Y DATOS DEL CONTACTO DEL CONTRATISTA:\n"
        "EMPRESA: Test SAS\n"
        "GERENTE DE PROYECTO: Juan Perez\n"
        "TELÉFONO: 3001234567\n"
        "NÚMERO DE PEDIDO: 12345"
    )
    hoja["C7"] = "FECHA DE ELABORACIÓN: 15/03/2026\n\nDURACIÓN (EN DÍAS): 5"
    hoja["H7"] = "DESCRIBA EL TRABAJO A REALIZAR: Trabajo de prueba"
    hoja["A14"] = "1. Actividad de prueba"
    hoja["B14"] = "Técnica de prueba"
    hoja["C14"] = "Riesgo de prueba"
    hoja["D14"], hoja["E14"], hoja["F14"] = 3, 3, 3
    hoja["H14"] = "Medidas de prueba"
    hoja["I14"], hoja["J14"], hoja["K14"] = 1, 1, 1
    hoja["M14"] = "Trabajos en Altura > 1.8 m"
    hoja["N14"] = "SI"

    hoja_fpe = libro.create_sheet("Firmas,Permisos, EPP")
    hoja_fpe["I4"] = "Trabajos en Altura > 1.8 m"
    hoja_fpe["K4"] = "X"
    hoja_fpe["L4"] = "Casco de seguridad"
    hoja_fpe["N4"] = "X"

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer.read()


class ImportarExcelDeclaracionTests(ApiTestsBase):
    def _archivo(self, contenido=None):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(
            "declaracion.xlsx",
            contenido if contenido is not None else _construir_excel_declaracion_prueba(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_requiere_autenticacion(self):
        response = self.client.post(
            reverse("contratistas:declaraciones_importar_excel"), {"archivo": self._archivo()}
        )
        self.assertEqual(response.status_code, 401)

    def test_parsea_encabezado_y_actividades(self):
        response = self.client.post(
            reverse("contratistas:declaraciones_importar_excel"),
            {"archivo": self._archivo()},
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["planta_area"], "Planta Prueba")
        self.assertEqual(response.data["numero_pedido"], "12345")
        self.assertEqual(response.data["gerente_proyecto"], "Juan Perez")
        self.assertEqual(response.data["contacto_telefono"], "3001234567")
        self.assertEqual(response.data["fecha_elaboracion"], "2026-03-15")
        self.assertEqual(response.data["duracion_dias"], 5)
        self.assertEqual(response.data["descripcion_trabajo"], "Trabajo de prueba")
        self.assertEqual(len(response.data["actividades"]), 1)
        actividad = response.data["actividades"][0]
        self.assertEqual(actividad["secuencia"], "1. Actividad de prueba")
        self.assertEqual(actividad["descripcion_riesgo"], "Riesgo de prueba")
        self.assertEqual(actividad["permisos_requeridos"], ["Trabajos en Altura > 1.8 m"])
        self.assertEqual(actividad["epp_requerido"], ["Casco de seguridad"])
        self.assertTrue(actividad["tarea_sif"])
        self.assertEqual(response.data["avisos"], [])

    def test_portal_contratista_tambien_puede_importar(self):
        portal_user = Usuario.objects.create_user("portal_import_test", "portal_import@x.com", "clave12345")
        portal_user.perfil.rol = PerfilUsuario.Rol.CONTRATISTA
        portal_user.perfil.contratista = self.contratista
        portal_user.perfil.save(update_fields=["rol", "contratista"])
        response = self.client.post(
            reverse("contratistas:declaraciones_importar_excel"),
            {"archivo": self._archivo()},
            **self._auth(portal_user),
        )
        self.assertEqual(response.status_code, 200, response.data)

    def test_hoja_faltante_devuelve_400(self):
        from openpyxl import Workbook
        import io

        libro = Workbook()
        libro.active.title = "Otra hoja"
        buffer = io.BytesIO()
        libro.save(buffer)
        buffer.seek(0)
        response = self.client.post(
            reverse("contratistas:declaraciones_importar_excel"),
            {"archivo": self._archivo(buffer.read())},
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("Declaración de Método", response.data["detail"])

    def test_extension_invalida_devuelve_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        archivo = SimpleUploadedFile("declaracion.txt", b"no es un excel", content_type="text/plain")
        response = self.client.post(
            reverse("contratistas:declaraciones_importar_excel"),
            {"archivo": archivo},
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)


class ArchivoOrigenDeclaracionTests(ApiTestsBase):
    """Adjuntar el Excel original a una declaración ya creada — para que
    quien revisa pueda abrirlo y compararlo contra lo que quedó cargado."""

    def _declaracion(self, contratista=None):
        return DeclaracionMetodo.objects.create(
            contratista=contratista or self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )

    def _archivo(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        return SimpleUploadedFile(
            "original.xlsx",
            _construir_excel_declaracion_prueba(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_requiere_autenticacion(self):
        declaracion = self._declaracion()
        url = reverse("contratistas:declaraciones_archivo_origen", args=[declaracion.pk])
        response = self.client.post(url, {"archivo": self._archivo()})
        self.assertEqual(response.status_code, 401)

    def test_sube_y_queda_disponible_en_el_detalle(self):
        declaracion = self._declaracion()
        url = reverse("contratistas:declaraciones_archivo_origen", args=[declaracion.pk])
        response = self.client.post(url, {"archivo": self._archivo()}, **self._auth(self.operador))
        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(response.data["archivo_origen_excel"])

        detalle = self.client.get(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]), **self._auth(self.operador)
        )
        self.assertTrue(detalle.data["archivo_origen_excel"])

    def test_portal_contratista_puede_subir_a_su_propia_declaracion(self):
        declaracion = self._declaracion()
        portal_user = Usuario.objects.create_user("portal_origen_test", "portal_origen@x.com", "clave12345")
        portal_user.perfil.rol = PerfilUsuario.Rol.CONTRATISTA
        portal_user.perfil.contratista = self.contratista
        portal_user.perfil.save(update_fields=["rol", "contratista"])
        url = reverse("contratistas:declaraciones_archivo_origen", args=[declaracion.pk])
        response = self.client.post(url, {"archivo": self._archivo()}, **self._auth(portal_user))
        self.assertEqual(response.status_code, 200, response.data)

    def test_portal_contratista_no_puede_subir_a_declaracion_de_otra_empresa(self):
        otra_empresa = EmpresaContratista.objects.create(empresa=self.empresa, nombre="Otra SAS", nit="900000000-1")
        declaracion_ajena = self._declaracion(contratista=otra_empresa)
        portal_user = Usuario.objects.create_user("portal_origen_test2", "portal_origen2@x.com", "clave12345")
        portal_user.perfil.rol = PerfilUsuario.Rol.CONTRATISTA
        portal_user.perfil.contratista = self.contratista
        portal_user.perfil.save(update_fields=["rol", "contratista"])
        url = reverse("contratistas:declaraciones_archivo_origen", args=[declaracion_ajena.pk])
        response = self.client.post(url, {"archivo": self._archivo()}, **self._auth(portal_user))
        self.assertEqual(response.status_code, 404)

    def test_extension_invalida_devuelve_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        declaracion = self._declaracion()
        archivo = SimpleUploadedFile("original.txt", b"no es un excel", content_type="text/plain")
        url = reverse("contratistas:declaraciones_archivo_origen", args=[declaracion.pk])
        response = self.client.post(url, {"archivo": archivo}, **self._auth(self.operador))
        self.assertEqual(response.status_code, 400)


class NotasAlertasTests(ApiTestsBase):
    """Notas del personal de SST/interventoría sobre una alerta puntual —
    identificada por código + orden de la actividad que la disparó."""

    def _declaracion(self):
        return DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )

    def test_requiere_personal_interno(self):
        declaracion = self._declaracion()
        url = reverse("contratistas:declaraciones_notas_alertas", args=[declaracion.pk])

        portal_user = Usuario.objects.create_user("portal_notas_test", "portal_notas@x.com", "clave12345")
        portal_user.perfil.rol = PerfilUsuario.Rol.CONTRATISTA
        portal_user.perfil.contratista = self.contratista
        portal_user.perfil.save(update_fields=["rol", "contratista"])

        response = self.client.get(url, **self._auth(portal_user))
        self.assertEqual(response.status_code, 403)

    def test_lista_vacia_al_principio(self):
        declaracion = self._declaracion()
        url = reverse("contratistas:declaraciones_notas_alertas", args=[declaracion.pk])
        response = self.client.get(url, **self._auth(self.operador))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, [])

    def test_crear_nota_y_listarla(self):
        declaracion = self._declaracion()
        url = reverse("contratistas:declaraciones_notas_alertas", args=[declaracion.pk])
        response = self.client.post(
            url,
            {
                "codigo_alerta": "altura_sin_epp_caida",
                "actividad_orden": 0,
                "texto": "Se validó en sitio con el supervisor, queda pendiente actualizar el EPP.",
            },
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["autor_nombre"], "operador1")

        listado = self.client.get(url, **self._auth(self.operador))
        self.assertEqual(len(listado.data), 1)
        self.assertEqual(listado.data[0]["codigo_alerta"], "altura_sin_epp_caida")

    def test_texto_vacio_devuelve_400(self):
        declaracion = self._declaracion()
        url = reverse("contratistas:declaraciones_notas_alertas", args=[declaracion.pk])
        response = self.client.post(
            url,
            {"codigo_alerta": "altura_sin_epp_caida", "actividad_orden": 0, "texto": "   "},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)


class AutorizacionIngresoTests(ApiTestsBase):
    def setUp(self):
        super().setUp()
        self.trabajador2 = Trabajador.objects.create(
            contratista=self.contratista, nombres="Laura", apellidos="Pinzón", documento="99887766"
        )

    def _payload(self, **extra):
        payload = {
            "contratista": self.contratista.pk,
            "fecha_inicio": "2026-08-01",
            "fecha_fin": "2026-08-31",
            "hora_inicio": "07:00",
            "hora_fin": "17:00",
            "area_trabajo": "Envasado — línea 3",
            "sitio_encuentro_emergencia": "Punto de encuentro Portería 2",
            "responsable_siso_nombre": "Carlos Pardo",
            "responsable_siso_telefono": "3001234567",
            "trabajadores": [
                {"trabajador": self.trabajador.pk, "incluido": True},
                {"trabajador": self.trabajador2.pk, "incluido": False, "motivo_exclusion": "Curso vencido"},
            ],
        }
        payload.update(extra)
        return payload

    def test_crear_con_inclusiones_y_exclusiones(self):
        response = self.client.post(
            reverse("contratistas:autorizaciones_ingreso_lista"),
            self._payload(),
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(len(response.data["trabajadores"]), 2)
        autorizacion = AutorizacionIngreso.objects.get(pk=response.data["id"])
        self.assertEqual(autorizacion.trabajadores.filter(incluido=True).count(), 1)
        self.assertEqual(autorizacion.trabajadores.filter(incluido=False).count(), 1)

    def test_excluir_sin_motivo_devuelve_400(self):
        payload = self._payload(
            trabajadores=[{"trabajador": self.trabajador.pk, "incluido": False, "motivo_exclusion": ""}]
        )
        response = self.client.post(
            reverse("contratistas:autorizaciones_ingreso_lista"),
            payload,
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)

    def test_fecha_fin_anterior_a_inicio_devuelve_400(self):
        payload = self._payload(fecha_inicio="2026-08-31", fecha_fin="2026-08-01")
        response = self.client.post(
            reverse("contratistas:autorizaciones_ingreso_lista"),
            payload,
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("fecha_fin", response.data)

    def test_vigente_segun_fechas(self):
        hoy = timezone.localdate()
        vigente = AutorizacionIngreso.objects.create(
            contratista=self.contratista,
            fecha_inicio=hoy - datetime.timedelta(days=1),
            fecha_fin=hoy + datetime.timedelta(days=1),
            area_trabajo="Bodega",
            responsable_siso_nombre="Ana",
        )
        vencida = AutorizacionIngreso.objects.create(
            contratista=self.contratista,
            fecha_inicio=hoy - datetime.timedelta(days=10),
            fecha_fin=hoy - datetime.timedelta(days=5),
            area_trabajo="Bodega",
            responsable_siso_nombre="Ana",
        )
        self.assertTrue(vigente.vigente)
        self.assertFalse(vencida.vigente)

    def test_editar_reemplaza_lista_de_trabajadores(self):
        response = self.client.post(
            reverse("contratistas:autorizaciones_ingreso_lista"),
            self._payload(),
            content_type="application/json",
            **self._auth(self.operador),
        )
        autorizacion_id = response.data["id"]

        response = self.client.patch(
            reverse("contratistas:autorizaciones_ingreso_detalle", args=[autorizacion_id]),
            {"trabajadores": [{"trabajador": self.trabajador.pk, "incluido": True}]},
            content_type="application/json",
            **self._auth(self.operador),
        )
        self.assertEqual(response.status_code, 200, response.data)
        autorizacion = AutorizacionIngreso.objects.get(pk=autorizacion_id)
        self.assertEqual(autorizacion.trabajadores.count(), 1)

    def test_operador_no_puede_eliminar(self):
        response = self.client.post(
            reverse("contratistas:autorizaciones_ingreso_lista"),
            self._payload(),
            content_type="application/json",
            **self._auth(self.operador),
        )
        url = reverse("contratistas:autorizaciones_ingreso_detalle", args=[response.data["id"]])
        respuesta = self.client.delete(url, **self._auth(self.operador))
        self.assertEqual(respuesta.status_code, 403)

    def test_crear_queda_registrado_en_auditoria(self):
        self.client.post(
            reverse("contratistas:autorizaciones_ingreso_lista"),
            self._payload(),
            content_type="application/json",
            **self._auth(self.operador),
        )
        registro = RegistroAuditoria.objects.filter(modelo="AutorizacionIngreso", accion="creado").latest("fecha")
        self.assertEqual(registro.usuario, self.operador)

    def test_descargar_pdf_con_inclusiones_y_exclusiones(self):
        response = self.client.post(
            reverse("contratistas:autorizaciones_ingreso_lista"),
            self._payload(responsable_siso_cargo="Coordinadora SISO"),
            content_type="application/json",
            **self._auth(self.operador),
        )
        autorizacion_id = response.data["id"]
        url = reverse("contratistas:autorizaciones_ingreso_pdf", args=[autorizacion_id])
        respuesta = self.client.get(url, **self._auth(self.operador))
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta["Content-Type"], "application/pdf")
        self.assertGreater(len(respuesta.content), 500)


class PortalContratistaTests(ApiTestsBase):
    """El rol Contratista ve y opera solo dentro de su propia empresa: lee
    su contratista/trabajadores/radicaciones/autorizaciones, pero solo
    Declaración de Método le permite escribir — y únicamente para enviar o
    subsanar, nunca para aprobarse o rechazarse a sí mismo."""

    def setUp(self):
        super().setUp()
        self.otro_contratista = EmpresaContratista.objects.create(empresa=self.empresa, nombre="OTRA SAS")
        self.otro_trabajador = Trabajador.objects.create(
            contratista=self.otro_contratista, nombres="Pedro", apellidos="Ruiz", documento="1122334455"
        )
        self.portal_user = Usuario.objects.create_user("portal_scepsa", "portal@scepsa.com", "clave12345")
        self.portal_user.perfil.rol = PerfilUsuario.Rol.CONTRATISTA
        self.portal_user.perfil.contratista = self.contratista
        self.portal_user.perfil.save(update_fields=["rol", "contratista"])

    def test_lista_empresas_solo_muestra_la_propia(self):
        response = self.client.get(
            reverse("contratistas:empresas_lista"), **self._auth(self.portal_user)
        )
        self.assertEqual(response.status_code, 200)
        ids = [fila["id"] for fila in response.data]
        self.assertEqual(ids, [self.contratista.pk])

    def test_no_puede_ver_trabajador_de_otra_empresa(self):
        url = reverse("contratistas:trabajadores_detalle", args=[self.otro_trabajador.pk])
        response = self.client.get(url, **self._auth(self.portal_user))
        self.assertEqual(response.status_code, 404)

    def test_lista_trabajadores_ignora_filtro_contratista_ajeno(self):
        response = self.client.get(
            reverse("contratistas:trabajadores_lista"),
            {"contratista": self.otro_contratista.pk},
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 200)
        ids = [fila["id"] for fila in response.data]
        self.assertEqual(ids, [self.trabajador.pk])

    def test_no_puede_crear_trabajador(self):
        response = self.client.post(
            reverse("contratistas:trabajadores_lista"),
            {"contratista": self.contratista.pk, "nombres": "X", "apellidos": "Y", "documento": "1"},
            content_type="application/json",
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 403)

    def test_no_puede_aprobar_radicacion(self):
        radicacion = RadicacionSeguridadSocial.objects.create(trabajador=self.trabajador, anio=2026, mes="ENERO")
        response = self.client.post(
            reverse("contratistas:radicaciones_aprobar", args=[radicacion.pk]),
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 403)

    def test_no_puede_ver_indicadores_dashboard(self):
        response = self.client.get(
            reverse("contratistas:indicadores_dashboard"), **self._auth(self.portal_user)
        )
        self.assertEqual(response.status_code, 403)

    def test_no_puede_gestionar_funcionarios(self):
        response = self.client.get(reverse("contratistas:funcionarios_lista"), **self._auth(self.portal_user))
        self.assertEqual(response.status_code, 403)

    def test_puede_crear_declaracion_enviada_para_su_empresa(self):
        payload = {
            "contratista": self.otro_contratista.pk,  # intenta suplantar otra empresa
            "planta_area": "Tapas",
            "fecha_elaboracion": "2026-07-11",
            "duracion_dias": 10,
            "descripcion_trabajo": "Mantenimiento",
            "estado": "enviada",
        }
        response = self.client.post(
            reverse("contratistas:declaraciones_lista"),
            payload,
            content_type="application/json",
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["contratista"], self.contratista.pk)
        self.assertEqual(response.data["estado"], "enviada")

    def test_no_puede_crear_declaracion_ya_rechazada(self):
        """"aprobada" nunca es alcanzable al crear (exige firmas, que no
        existen aún) — la transición que sí hay que bloquear explícitamente
        es "rechazada", que solo exige un motivo."""
        payload = {
            "contratista": self.contratista.pk,
            "planta_area": "Tapas",
            "fecha_elaboracion": "2026-07-11",
            "duracion_dias": 10,
            "descripcion_trabajo": "Mantenimiento",
            "estado": "rechazada",
            "observaciones": "No cumple.",
        }
        response = self.client.post(
            reverse("contratistas:declaraciones_lista"),
            payload,
            content_type="application/json",
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 403)

    def test_no_puede_aprobarse_a_si_mismo_al_editar(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        FirmaMetodo.objects.create(
            declaracion=declaracion, rol="supervisor_contratista", nombre_firmante="Ana", firmante_usuario=self.admin
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"estado": "aprobada"},
            content_type="application/json",
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 403)
        declaracion.refresh_from_db()
        self.assertEqual(declaracion.estado, "borrador")

    def test_puede_subsanar_y_reenviar_declaracion_rechazada(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
            estado=DeclaracionMetodo.Estado.RECHAZADA,
            observaciones="Falta el permiso de trabajo en altura.",
        )
        response = self.client.patch(
            reverse("contratistas:declaraciones_detalle", args=[declaracion.pk]),
            {"descripcion_trabajo": "Instalación de pórtico con permiso adjunto", "estado": "enviada"},
            content_type="application/json",
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["estado"], "enviada")

    def test_no_ve_declaracion_de_otra_empresa(self):
        declaracion_ajena = DeclaracionMetodo.objects.create(
            contratista=self.otro_contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Otra obra",
        )
        response = self.client.get(
            reverse("contratistas:declaraciones_detalle", args=[declaracion_ajena.pk]),
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 404)

    def test_solo_puede_firmar_como_supervisor_contratista(self):
        declaracion = DeclaracionMetodo.objects.create(
            contratista=self.contratista,
            fecha_elaboracion=datetime.date(2026, 7, 11),
            descripcion_trabajo="Instalación de pórtico",
        )
        response = self.client.post(
            reverse("contratistas:declaraciones_firmar", args=[declaracion.pk]),
            {"rol": "delegado_abi", "nombre_firmante": "Alguien", "consiento_firma": True},
            content_type="application/json",
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 403)

        response = self.client.post(
            reverse("contratistas:declaraciones_firmar", args=[declaracion.pk]),
            {"rol": "supervisor_contratista", "nombre_firmante": "Ana", "consiento_firma": True},
            content_type="application/json",
            **self._auth(self.portal_user),
        )
        self.assertEqual(response.status_code, 201, response.data)
